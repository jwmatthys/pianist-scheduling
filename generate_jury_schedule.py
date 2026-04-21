"""
Jury Day Schedule Generator

Reads:
  - JuryList .xlsx, which must contain:
      "Pianist" sheet       — one row per student: Name, Area, Instrument,
                              Pianist (last name; blank = no pianist needed),
                              Jury? (Y to include)
      "Jury Information"    — Area, Jury Length, Start Time, Hourly Break,
                              Lunch Break, Location
      "Pianist - <Name>"    — one sheet per accompanist; availability is read
                              from the "Jury Day Availability" column (col 8).
                              Non-"Available" slots (including Tentative) are
                              treated as unavailable.

Writes:
  - jury_schedule_<timestamp>.xlsx  (same directory as input)

Scheduling rules:
  - Assigned pianists are inflexible (no substitutions).
  - Start times in Jury Information are the EARLIEST possible, not fixed.
  - Areas sharing a room are scheduled sequentially with no overlap.
  - Within an area, students are pre-sorted into pianist blocks ordered by
    when each pianist finishes their other-area commitments, minimising idle
    gaps across the full day.
  - Gaps within an area are minimised by jumping to the next free minute.
  - Hourly breaks (10 min) and lunch breaks (30 min) are inserted when flagged.
  - 'Tentative' availability is treated as unavailable.
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')


# ── Helpers ───────────────────────────────────────────────────────────────────

def to_minutes(val):
    """Convert timedelta or time to integer minutes from midnight."""
    if isinstance(val, timedelta):
        return int(val.total_seconds() // 60)
    return val.hour * 60 + val.minute

def fmt(minutes):
    h, m = divmod(int(minutes), 60)
    suffix = 'AM' if h < 12 else 'PM'
    h12 = h % 12 or 12
    return f"{h12}:{m:02d} {suffix}"


# ── Data loading ──────────────────────────────────────────────────────────────

def load_pianist_unavailability(xl):
    """
    Returns {pianist_name: set_of_unavailable_minutes}.
    Reads from the "Jury Day Availability" column (column 8, index 7) of each
    "Pianist - <Name>" sheet. Non-"Available" slots (including Tentative) mark
    all 30 minutes of that slot as unavailable.
    """
    result = {}
    for sheet in [s for s in xl.sheet_names if s.startswith('Pianist -')]:
        df = pd.read_excel(xl, sheet_name=sheet, header=None)
        name = str(df.iloc[0, 1]).strip()
        unavail = set()
        for _, row in df.iloc[3:].iterrows():
            if pd.isna(row.iloc[0]):
                break
            val = row.iloc[7] if len(row) > 7 and not pd.isna(row.iloc[7]) else ''
            if str(val).strip() != 'Available':
                base = to_minutes(row.iloc[0])
                unavail.update(range(base, base + 30))
        result[name] = unavail
    return result

def load_students(xl):
    """
    Read the Pianist sheet.
    Returns {area: [(name, instrument, needs_pianist, pianist), ...]}.
    Rows where Jury? != 'Y' are skipped.
    A blank Pianist cell means the student does not need an accompanist.
    """
    df = pd.read_excel(xl, sheet_name='Pianist')
    area_students = {}
    for _, row in df.iterrows():
        if str(row.get('Jury?', '')).strip().upper() != 'Y':
            continue
        name    = f"{str(row['Student Preferred']).strip()} {str(row['Student Last Name']).strip()}"
        area    = str(row['Area']).strip()
        instr   = str(row.get('Instrument', '')).strip() if pd.notna(row.get('Instrument')) else ''
        pianist = str(row['Pianist']).strip() if pd.notna(row.get('Pianist')) else ''
        needs   = bool(pianist)
        area_students.setdefault(area, []).append((name, instr, needs, pianist))
    return area_students


# ── Pianist booking ───────────────────────────────────────────────────────────

def is_free(pianist_unavail, pianist_booked, name, start, duration):
    """True if pianist is free for the entire [start, start+duration) window."""
    busy = pianist_unavail.get(name, set()) | pianist_booked.get(name, set())
    return not any((start + k) in busy for k in range(duration))

def book(pianist_booked, name, start, duration):
    pianist_booked.setdefault(name, set()).update(range(start, start + duration))

def next_free(pianist_unavail, pianist_booked, name, from_min, duration):
    """Earliest minute >= from_min where pianist is free for 'duration' minutes."""
    t = from_min
    while t < 23 * 60:
        if is_free(pianist_unavail, pianist_booked, name, t, duration):
            return t
        t += 1
    return None


# ── Area scheduling ───────────────────────────────────────────────────────────

def find_actual_start(students, earliest, slot_min, pianist_unavail, pianist_booked):
    """
    Return the best start time >= earliest for this area.

    The goal is to eliminate avoidable gaps while not starting earlier than
    necessary.  The algorithm works as follows:

    1. Find the naive earliest start: the first moment at least one student
       can be placed.

    2. Identify any 'unavailability boundary': the earliest time a currently-
       blocked pianist becomes free.  Students placeable before that boundary
       will create a gap if scheduled too early.

    3. If placing those students starting at the naive start would leave a gap
       larger than one slot, delay the start so they fill the time immediately
       before the boundary instead — eliminating the gap.

    This handles cases like: two no-pianist students + Roberts available from
    09:00, but Matthys/Schaner not free until 10:30.  Rather than starting at
    09:00 and leaving a 42-minute gap, we start at 10:04 so the available
    students run flush into 10:30.
    """
    if not students:
        return earliest

    # Step 1: naive earliest start — first moment any student can be placed
    naive = earliest
    while naive < 23 * 60:
        for _, _, needs, pianist in students:
            free = is_free(pianist_unavail, pianist_booked, pianist, naive, slot_min) \
                   if (needs and pianist) else True
            if free:
                break
        else:
            # nothing placeable — jump to next free transition
            nf_times = [
                next_free(pianist_unavail, pianist_booked, p, naive + 1, slot_min)
                for _, _, needs, p in students if needs and p
            ]
            nf_times = [t for t in nf_times if t is not None]
            naive = min(nf_times) if nf_times else naive + 1
            continue
        break

    # Step 2: find the earliest unavailability boundary — the first time a
    # currently-blocked pianist becomes free (after naive)
    boundary = None
    for _, _, needs, pianist in students:
        if not (needs and pianist):
            continue
        if not is_free(pianist_unavail, pianist_booked, pianist, naive, slot_min):
            nf = next_free(pianist_unavail, pianist_booked, pianist, naive, slot_min)
            if nf is not None:
                boundary = nf if boundary is None else min(boundary, nf)

    if boundary is None:
        return naive  # no blocked pianists — start as early as possible

    # Step 3: greedy-simulate from naive up to boundary, counting how many
    # students can be placed. Uses the same "pick any free student" logic as
    # the real greedy, so ordering in the original sheet doesn't matter —
    # Kwiecien/Roberts students get counted even if Schaner is first on the sheet.
    sim_remaining = list(students)
    sim_booked = {k: set(v) for k, v in pianist_booked.items()}
    sim_t = naive
    pre_boundary = 0
    while sim_remaining and sim_t < boundary:
        for i, (_, _, needs, pianist) in enumerate(sim_remaining):
            free = is_free(pianist_unavail, sim_booked, pianist, sim_t, slot_min) \
                   if (needs and pianist) else True
            if free:
                if needs and pianist:
                    sim_booked.setdefault(pianist, set()).update(range(sim_t, sim_t + slot_min))
                sim_remaining.pop(i)
                pre_boundary += 1
                sim_t += slot_min
                break
        else:
            break  # nothing free at this slot — stop counting

    if pre_boundary == 0:
        return naive  # nothing placeable before boundary anyway

    # Find the earliest start >= naive such that the resulting gap to the boundary
    # is at most 3 slot-lengths (matching the hourly-break cycle).  Starting
    # earlier than this would leave an uncomfortably long gap; starting later
    # wastes time needlessly.
    #
    # Example: 3 available students, boundary=10:30, slot=10min.
    #   naive=9:00 → gap=60m (too large). delayed=10:00 → gap=0m but 60m late.
    #   Best: start=9:30 → gap=30m (≤3 slots), delay=30m — earliest that fits.
    max_gap = 3 * slot_min
    gap_if_naive = boundary - sim_t  # gap if we start at naive
    if gap_if_naive > max_gap:
        # Shift start forward until the gap ≤ max_gap
        delayed = boundary - pre_boundary * slot_min - max_gap
        return max(naive, delayed)

    return naive


def schedule_area(students, earliest, slot_min, hourly_break, lunch_break,
                  pianist_unavail, pianist_booked, cross_area_booked):
    """
    Schedule students within one area.

    'students' is a list of (name, instrument, needs_pianist, pianist).
    'pianist_booked' is shared across all areas and updated in place.
    'cross_area_booked' is a snapshot of pianist_booked before this area started,
    representing only commitments from OTHER areas (used for sorting, not conflict checks).

    Strategy:
    - Find the actual start time (find_actual_start): the earliest moment at
      least one student can be placed. If some pianists are unavailable, we
      start with whoever is ready and accept the resulting gap — it would
      exist regardless of start time.
    - Pre-sort remaining students to form compact pianist blocks:
        1. Group students by pianist.
        2. Order groups by when that pianist last appears in cross_area_booked
           (i.e. when they finish their other-area work). Pianists who finish
           earlier get scheduled earlier in this area, so each pianist's slots
           are as contiguous as possible across the whole day.
        3. Within each group the original sheet order is preserved.
    - Greedily place students in that order. If the next student's pianist is
      temporarily unavailable (cross-area conflict), skip forward to the next
      student whose pianist IS free, then return to deferred students once the
      blocking conflict resolves.
    - Jump forward in time (not minute-by-minute) when no student can be placed.
    - Insert hourly (10-min) and lunch (30-min) breaks as configured.

    Returns (slots, end_minute).
    """
    if not students:
        return [], earliest

    actual_start = find_actual_start(students, earliest, slot_min,
                                     pianist_unavail, pianist_booked)

    # ── Pre-sort students into compact pianist blocks ─────────────────────────
    # For each pianist, find their last minute of cross-area commitment.
    # Groups are ordered so pianists who finish their other work earliest
    # are scheduled first — minimising idle gaps across the full day.
    def first_available_minute(pianist):
        """Earliest minute >= actual_start when this pianist is free for one slot.
        Pianists available right now return actual_start; blocked ones return
        when they next become free. This ensures available pianists always sort
        before unavailable ones, filling the pre-boundary window cleanly."""
        if not pianist:
            return actual_start
        t = actual_start
        while t < 23 * 60:
            if is_free(pianist_unavail, pianist_booked, pianist, t, slot_min):
                return t
            t += 1
        return actual_start

    def last_cross_area_minute(pianist):
        """Latest minute this pianist is committed in other areas (or 0 if none)."""
        busy = pianist_unavail.get(pianist, set()) | cross_area_booked.get(pianist, set())
        return max(busy) if busy else 0

    # Group students by pianist, preserving original within-group order
    groups = {}
    group_order = []
    for student in students:
        pianist = student[3]  # (name, instrument, needs_pianist, pianist)
        if pianist not in groups:
            groups[pianist] = []
            group_order.append(pianist)
        groups[pianist].append(student)

    # Sort groups: pianists available soonest go first (fills pre-boundary window
    # cleanly); break ties by last cross-area commitment (earlier = first).
    group_order.sort(key=lambda p: (first_available_minute(p), last_cross_area_minute(p)) if p else (-1, -1))

    # Flatten back into a sorted student list
    remaining = []
    for pianist in group_order:
        remaining.extend(groups[pianist])
    # ─────────────────────────────────────────────────────────────────────────

    t = actual_start
    area_start = t
    slots = []
    juries_since_reset = 0
    juries_per_cycle = (60 - 10) // slot_min
    lunch_done = False

    while remaining:
        # Insert 10-min break after every full cycle of juries,
        # but skip it if only one student remains (no point breaking for one).
        if hourly_break and juries_since_reset > 0 and juries_since_reset % juries_per_cycle == 0 \
                and len(remaining) > 1:
            slots.append({'type': 'break', 'time': t, 'label': '10-Minute Break'})
            t += 10
            juries_since_reset = 0

        # 30-min lunch break when we reach noon; reset jury count
        if lunch_break and not lunch_done and t >= 12 * 60:
            lunch_done = True
            slots.append({'type': 'lunch', 'time': t, 'label': 'Lunch Break (30 min)'})
            t += 30
            juries_since_reset = 0

        # Place the earliest student in the sorted list whose pianist is free now.
        # This preserves the block ordering while skipping temporarily blocked pianists.
        placed = False
        for i, (name, instrument, needs_pianist, pianist) in enumerate(remaining):
            free = is_free(pianist_unavail, pianist_booked, pianist, t, slot_min) \
                   if (needs_pianist and pianist) else True
            if free:
                if needs_pianist and pianist:
                    book(pianist_booked, pianist, t, slot_min)
                slots.append({
                    'type':         'student',
                    'time':         t,
                    'student':      name,
                    'instrument':   instrument,
                    'need_pianist': needs_pianist,
                    'pianist':      pianist if needs_pianist else '',
                    'duration':     slot_min,
                })
                remaining.pop(i)
                t += slot_min
                juries_since_reset += 1
                placed = True
                break

        if not placed:
            # Jump directly to the next minute any remaining student's pianist is free
            candidates = []
            for name, instrument, needs_pianist, pianist in remaining:
                if needs_pianist and pianist:
                    nf = next_free(pianist_unavail, pianist_booked, pianist, t + 1, slot_min)
                    if nf is not None:
                        candidates.append(nf)
                else:
                    candidates.append(t + 1)
            t = min(candidates) if candidates else t + 1

    return slots, t



# ── Full schedule ─────────────────────────────────────────────────────────────

def build_schedule(area_students, jury_info, pianist_unavail):

    # Parse jury info
    area_info = {}
    for _, row in jury_info.iterrows():
        area_info[row['Area']] = {
            'earliest':     to_minutes(row['Start Time']),
            'slot_min':     int(row['Jury Length']),
            'hourly_break': bool(row['Hourly Break']),
            'lunch_break':  bool(row['Lunch Break']),
            'location':     row['Location'],
        }

    # Group areas by room and sort by earliest start time
    rooms = {}
    for area, info in area_info.items():
        rooms.setdefault(info['location'], []).append(area)
    for room in rooms:
        rooms[room].sort(key=lambda a: area_info[a]['earliest'])

    # pianist_booked is shared across ALL areas (cross-area conflict tracking)
    pianist_booked = {}
    area_schedule = {}

    for room, area_list in rooms.items():
        room_cursor = None

        for area in area_list:
            info = area_info[area]
            earliest = info['earliest'] if room_cursor is None \
                       else max(info['earliest'], room_cursor)
            students = area_students.get(area, [])

            # Snapshot booked state before this area — used for cross-area urgency sorting
            cross_area_booked = {k: set(v) for k, v in pianist_booked.items()}
            slots, end = schedule_area(
                students, earliest,
                info['slot_min'], info['hourly_break'], info['lunch_break'],
                pianist_unavail, pianist_booked, cross_area_booked,
            )

            actual_start = slots[0]['time'] if slots else earliest
            area_schedule[area] = {
                'location':     room,
                'earliest':     info['earliest'],
                'actual_start': actual_start,
                'actual_end':   end,
                'slots':        slots,
                'slot_min':     info['slot_min'],
            }
            room_cursor = end

    return area_schedule


# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(area_schedule, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    DARK_BLUE  = '1F4E79'
    MID_BLUE   = '2E75B6'
    LIGHT_BLUE = 'DDEEFF'
    YELLOW     = 'FFF2CC'
    AMBER      = 'FFE0A0'
    WHITE      = 'FFFFFF'

    thin = Side(style='thin', color='AAAAAA')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def cell(ws, r, col, val, bold=False, size=10, fgcolor='000000',
             bg=None, align=None, border=None):
        c = ws.cell(row=r, column=col, value=val)
        c.font  = Font(name='Arial', bold=bold, size=size, color=fgcolor)
        if bg:     c.fill      = PatternFill('solid', start_color=bg)
        if align:  c.alignment = align
        if border: c.border    = border
        return c

    ordered = sorted(area_schedule.items(),
                     key=lambda x: (x[1]['location'], x[1]['actual_start']))

    # ── Summary sheet ─────────────────────────────────────────────────
    ws = wb.create_sheet('Summary')
    for col, w in zip('ABCDEF', [15, 15, 10, 10, 12, 30]):
        ws.column_dimensions[col].width = w
    ws.merge_cells('A1:F1')
    cell(ws,1,1,'JURY DAY SCHEDULE — SUMMARY',
         bold=True, size=14, fgcolor=WHITE, bg=DARK_BLUE, align=ctr)
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(['Area','Location','Start','End','# Students','Notes'], 1):
        cell(ws,2,i,h, bold=True, size=10, fgcolor=WHITE, bg=MID_BLUE, align=ctr, border=bdr)

    prev_loc = None
    r = 3
    for area, sched in ordered:
        students = [s for s in sched['slots'] if s['type'] == 'student']
        # Compute internal gaps
        gaps, prev_end = [], None
        for slot in sched['slots']:
            if slot['type'] == 'break':
                prev_end = slot['time'] + 10; continue
            if slot['type'] == 'lunch':
                prev_end = slot['time'] + 30; continue
            if prev_end and slot['time'] > prev_end:
                gaps.append(slot['time'] - prev_end)
            prev_end = slot['time'] + slot['duration']

        bg = LIGHT_BLUE if sched['location'] != prev_loc else WHITE
        prev_loc = sched['location']

        notes = []
        delay = sched['actual_start'] - sched['earliest']
        if delay > 0:
            notes.append(f"Starts {delay}m after earliest ({fmt(sched['earliest'])})")
        if gaps:
            notes.append(f"{sum(gaps)}m wait gap(s) within area")

        cell(ws,r,1,area,                        bg=bg, align=lft, border=bdr)
        cell(ws,r,2,sched['location'],            bg=bg, align=ctr, border=bdr)
        cell(ws,r,3,fmt(sched['actual_start']),   bg=bg, align=ctr, border=bdr)
        cell(ws,r,4,fmt(sched['actual_end']),     bg=bg, align=ctr, border=bdr)
        cell(ws,r,5,len(students),                bg=bg, align=ctr, border=bdr)
        note_bg = AMBER if notes else bg
        cell(ws,r,6,'; '.join(notes), bg=note_bg, align=lft, border=bdr)
        r += 1

    # ── Per-area sheets ───────────────────────────────────────────────
    for area, sched in ordered:
        ws = wb.create_sheet(area)
        for col, w in zip('ABCDE', [10, 20, 16, 18, 14]):
            ws.column_dimensions[col].width = w
        ws.merge_cells('A1:E1')
        title = (f"{area} Jury  ·  {sched['location']}  ·  "
                 f"{fmt(sched['actual_start'])}–{fmt(sched['actual_end'])}"
                 f"  ({sched['slot_min']} min/student)")
        cell(ws,1,1,title, bold=True, size=12, fgcolor=WHITE, bg=DARK_BLUE, align=ctr)
        ws.row_dimensions[1].height = 26
        for i, h in enumerate(['Time','Student','Instrument','Pianist','Notes'], 1):
            cell(ws,2,i,h, bold=True, size=10, fgcolor=WHITE, bg=MID_BLUE, align=ctr, border=bdr)

        r, alt = 3, False
        prev_end = None
        for slot in sched['slots']:
            if slot['type'] == 'break':
                ws.merge_cells(f'A{r}:E{r}')
                cell(ws,r,1,f"— {slot['label']} ({fmt(slot['time'])}) —",
                     bold=True, bg=YELLOW, align=ctr, border=bdr)
                prev_end = slot['time'] + 10
                ws.row_dimensions[r].height = 18
                r += 1; alt = False; continue
            if slot['type'] == 'lunch':
                ws.merge_cells(f'A{r}:E{r}')
                cell(ws,r,1,f"— {slot['label']} ({fmt(slot['time'])}) —",
                     bold=True, bg=YELLOW, align=ctr, border=bdr)
                prev_end = slot['time'] + 30
                ws.row_dimensions[r].height = 18
                r += 1; alt = False; continue

            bg = LIGHT_BLUE if alt else WHITE
            alt = not alt
            note = ''
            if prev_end is not None and slot['time'] > prev_end:
                note = f"+{slot['time'] - prev_end}m wait"
                bg = AMBER

            cell(ws,r,1,fmt(slot['time']),  bg=bg, align=ctr, border=bdr)
            cell(ws,r,2,slot['student'],    bg=bg, align=lft, border=bdr)
            cell(ws,r,3,slot['instrument'], bg=bg, align=ctr, border=bdr)
            cell(ws,r,4,slot['pianist'],    bg=bg, align=ctr, border=bdr)
            cell(ws,r,5,note,               bg=bg, align=ctr, border=bdr)
            ws.row_dimensions[r].height = 18
            prev_end = slot['time'] + slot['duration']
            r += 1

    # ── Per-pianist sheets ────────────────────────────────────────────
    # Collect every student slot across all areas, keyed by pianist
    pianist_slots = {}
    for area, sched in area_schedule.items():
        for slot in sched['slots']:
            if slot['type'] != 'student' or not slot['pianist']:
                continue
            pianist_slots.setdefault(slot['pianist'], []).append({
                'time':       slot['time'],
                'end':        slot['time'] + slot['duration'],
                'student':    slot['student'],
                'instrument': slot['instrument'],
                'area':       area,
                'location':   sched['location'],
                'duration':   slot['duration'],
            })

    for pianist, bookings in sorted(pianist_slots.items()):
        bookings.sort(key=lambda s: s['time'])
        ws = wb.create_sheet(f'Pianist - {pianist}')
        for col, w in zip('ABCDEF', [10, 10, 20, 14, 14, 14]):
            ws.column_dimensions[col].width = w

        # Title
        first, last = bookings[0]['time'], bookings[-1]['end']
        total_busy  = sum(b['duration'] for b in bookings)
        total_idle  = last - first - total_busy
        ws.merge_cells('A1:F1')
        cell(ws,1,1,
             f"{pianist}  ·  {fmt(first)}–{fmt(last)}  ·  {len(bookings)} students  ·  {total_idle}m idle",
             bold=True, size=12, fgcolor=WHITE, bg=DARK_BLUE, align=ctr)
        ws.row_dimensions[1].height = 26

        for i, h in enumerate(['Start','End','Student','Instrument','Area','Location'], 1):
            cell(ws,2,i,h, bold=True, size=10, fgcolor=WHITE, bg=MID_BLUE, align=ctr, border=bdr)

        r, alt, prev_end = 3, False, None
        for b in bookings:
            # Gap row
            if prev_end is not None and b['time'] > prev_end:
                gap = b['time'] - prev_end
                ws.merge_cells(f'A{r}:F{r}')
                cell(ws,r,1, f"— {gap}m gap ({fmt(prev_end)}–{fmt(b['time'])}) —",
                     bold=True, bg=AMBER, align=ctr, border=bdr)
                ws.row_dimensions[r].height = 16
                r += 1
                alt = False

            bg = LIGHT_BLUE if alt else WHITE
            alt = not alt
            cell(ws,r,1, fmt(b['time']),    bg=bg, align=ctr, border=bdr)
            cell(ws,r,2, fmt(b['end']),     bg=bg, align=ctr, border=bdr)
            cell(ws,r,3, b['student'],      bg=bg, align=lft, border=bdr)
            cell(ws,r,4, b['instrument'],   bg=bg, align=ctr, border=bdr)
            cell(ws,r,5, b['area'],         bg=bg, align=ctr, border=bdr)
            cell(ws,r,6, b['location'],     bg=bg, align=ctr, border=bdr)
            ws.row_dimensions[r].height = 18
            prev_end = b['end']
            r += 1

    wb.save(output_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Generate jury day schedule.')
    parser.add_argument('--input', required=True,
                        help='JuryList .xlsx (Pianist sheet, Jury Information sheet, Pianist - * sheets)')
    args = parser.parse_args()

    timestamp   = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir  = os.path.dirname(os.path.abspath(args.input)) or '.'
    if not os.access(output_dir, os.W_OK):
        output_dir = '.'
    output_file = os.path.join(output_dir, f'jury_schedule_{timestamp}.xlsx')

    xl              = pd.ExcelFile(args.input)
    jury_info       = pd.read_excel(xl, sheet_name='Jury Information')
    pianist_unavail = load_pianist_unavailability(xl)
    area_students   = load_students(xl)

    area_schedule = build_schedule(area_students, jury_info, pianist_unavail)
    write_excel(area_schedule, output_file)
    print(f"Saved: {output_file}\n")

    print(f"{'Area':<12} {'Location':<14} {'Earliest':>8} {'Start':>6} {'End':>6}  {'#':>3}  Notes")
    print('─' * 75)
    for area, sched in sorted(area_schedule.items(),
                               key=lambda x: (x[1]['location'], x[1]['actual_start'])):
        students = [s for s in sched['slots'] if s['type'] == 'student']
        gaps, prev_end = [], None
        for slot in sched['slots']:
            if slot['type'] == 'break':  prev_end = slot['time'] + 10;  continue
            if slot['type'] == 'lunch':  prev_end = slot['time'] + 30;  continue
            if prev_end and slot['time'] > prev_end:
                gaps.append(slot['time'] - prev_end)
            prev_end = slot['time'] + slot['duration']
        notes = []
        delay = sched['actual_start'] - sched['earliest']
        if delay: notes.append(f"delayed +{delay}m")
        if gaps:  notes.append(f"{sum(gaps)}m gap")
        print(f"{area:<12} {sched['location']:<14}"
              f" {fmt(sched['earliest']):>8} {fmt(sched['actual_start']):>6}"
              f" {fmt(sched['actual_end']):>6}  {len(students):>3}  {'  '.join(notes)}")

if __name__ == '__main__':
    main()

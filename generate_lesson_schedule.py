"""
Lesson Schedule by Room Generator

Reads:
  - sample_lesson_schedule.xlsx (Lessons sheet)

Writes:
  - lesson_schedule_by_room_<timestamp>.xlsx

Output: Single sheet with all lessons organized by room, then by day and start time.
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

DAYS_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

DARK_BLUE  = '1F4E79'
MID_BLUE   = '2E75B6'
LIGHT_BLUE = 'DDEEFF'
ALT_BLUE   = 'EBF3FB'
WHITE      = 'FFFFFF'
YELLOW     = 'FFF9E6'
ROOM_HEADER_BG = '264478'

DAY_COLORS = {
    'Monday':    'E8F4FD',
    'Tuesday':   'FEF9E7',
    'Wednesday': 'E9F7EF',
    'Thursday':  'FDEDEC',
    'Friday':    'F4ECF7',
    'Saturday':  'FDF2E9',
    'Sunday':    'F2F3F4',
}


def fmt_time(val):
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        h, rem = divmod(total, 3600)
        m = rem // 60
        return f"{h:02d}:{m:02d}"
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M')
    return str(val)


def load_lessons(workbook_path):
    xl = pd.ExcelFile(workbook_path)
    df = pd.read_excel(xl, sheet_name='Lessons')
    # Drop blank rows and reference data at bottom
    df = df[df['Lesson Teacher Name'].notna() & df['Lesson Day'].notna()]
    df = df[df['Lesson Day'].isin(DAYS_ORDER)]
    df = df.reset_index(drop=True)
    return df


def normalize_location(loc):
    return str(loc).strip() if pd.notna(loc) else ''


def write_schedule(df, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule by Room'
    ws.sheet_view.showGridLines = False

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_bottom = Border(
        left=thin, right=thin,
        top=thin,
        bottom=Side(style='medium', color='888888')
    )
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def cell(ws, r, col, val, bold=False, size=10, fgcolor='000000',
             bg=None, align=None, border=None, italic=False):
        c = ws.cell(row=r, column=col, value=val)
        c.font  = Font(name='Arial', bold=bold, size=size, color=fgcolor, italic=italic)
        if bg:     c.fill      = PatternFill('solid', start_color=bg)
        if align:  c.alignment = align
        if border: c.border    = border
        return c

    # Column layout
    col_headers = ['Room', 'Day', 'Start', 'End', 'Teacher', 'Student', 'Instrument', 'Area', 'Pianist?', 'Required Pianist']
    col_keys    = ['Lesson Location', 'Lesson Day', 'Lesson Start Time', 'Lesson End Time',
                   'Lesson Teacher Name', 'Student Name', 'Instrument', 'Area', 'Need Pianist', 'Required Pianist']
    col_widths  = [10, 12, 8, 8, 16, 14, 14, 12, 10, 18]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ncols = len(col_headers)

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ts = datetime.now().strftime('%Y-%m-%d %H:%M')
    c = ws.cell(row=1, column=1, value=f'Lesson Schedule by Room   ·   Generated {ts}')
    c.font      = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill      = PatternFill('solid', start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Column headers
    for i, h in enumerate(col_headers, 1):
        cell(ws, 2, i, h, bold=True, size=10, fgcolor=WHITE, bg=MID_BLUE, align=ctr, border=bdr)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'

    # Sort data
    df = df.copy()
    df['_loc_norm'] = df['Lesson Location'].apply(normalize_location)
    df['_day_ord']  = df['Lesson Day'].map(lambda d: DAYS_ORDER.index(d) if d in DAYS_ORDER else 99)
    df['_start_min'] = df['Lesson Start Time'].apply(lambda t: (
        int(t.total_seconds() // 60) if isinstance(t, timedelta)
        else (t.hour * 60 + t.minute if hasattr(t, 'hour') else 0)
    ))
    df = df.sort_values(['_loc_norm', '_day_ord', '_start_min']).reset_index(drop=True)

    row = 3
    locations = df['_loc_norm'].unique()

    for loc in locations:
        loc_df = df[df['_loc_norm'] == loc]

        # Room header
        ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
        lesson_count = len(loc_df)
        c = ws.cell(row=row, column=1, value=f'Room: {loc}    ({lesson_count} lesson{"s" if lesson_count != 1 else ""})')
        c.font      = Font(name='Arial', bold=True, size=11, color=WHITE)
        c.fill      = PatternFill('solid', start_color=ROOM_HEADER_BG)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border    = bdr
        ws.row_dimensions[row].height = 20
        row += 1

        current_day = None
        alt = False

        for _, lesson in loc_df.iterrows():
            day = lesson['Lesson Day']

            # Day sub-header when day changes
            if day != current_day:
                current_day = day
                alt = False
                day_bg = DAY_COLORS.get(day, 'F5F5F5')
                ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
                c = ws.cell(row=row, column=1, value=f'  {day}')
                c.font      = Font(name='Arial', bold=True, size=10, color='1F4E79', italic=False)
                c.fill      = PatternFill('solid', start_color=day_bg)
                c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                c.border    = thick_bottom
                ws.row_dimensions[row].height = 18
                row += 1

            bg = ALT_BLUE if alt else WHITE
            alt = not alt

            for i, key in enumerate(col_keys, 1):
                val = lesson.get(key, '')
                # Format times
                if key in ('Lesson Start Time', 'Lesson End Time'):
                    val = fmt_time(val)
                # Format Need Pianist
                elif key == 'Need Pianist':
                    val = 'Yes' if str(val).strip() in ('1', '1.0', 'True', 'TRUE', 'Yes') else 'No'
                # Clean NaN
                elif pd.isna(val) if not isinstance(val, str) else False:
                    val = ''
                else:
                    val = str(val).strip() if not isinstance(val, (int, float)) else val

                align = ctr if key in ('Lesson Start Time', 'Lesson End Time', 'Need Pianist', 'Lesson Day', 'Lesson Location') else lft
                cell(ws, row, i, val, bg=bg, align=align, border=bdr)

            ws.row_dimensions[row].height = 17
            row += 1

        # Spacer row between rooms
        for i in range(1, ncols + 1):
            ws.cell(row=row, column=i).fill = PatternFill('solid', start_color='F0F0F0')
        ws.row_dimensions[row].height = 6
        row += 1

    # Summary at bottom
    row += 1
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1,
                value=f'Total: {len(df)} lessons across {len(locations)} rooms')
    c.font      = Font(name='Arial', bold=True, size=10, color='595959', italic=True)
    c.alignment = Alignment(horizontal='right', vertical='center')

    wb.save(output_path)
    print(f'Saved: {output_path}')
    print(f'  {len(df)} lessons | {len(locations)} rooms')


def main():
    parser = argparse.ArgumentParser(description='Generate lesson schedule by room.')
    parser.add_argument('--workbook', required=True,
                        help='Excel workbook containing the Lessons sheet')
    args = parser.parse_args()

    timestamp   = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir  = os.path.dirname(os.path.abspath(args.workbook))
    output_file = os.path.join(output_dir, f'lesson_schedule_by_room_{timestamp}.xlsx')

    df = load_lessons(args.workbook)
    write_schedule(df, output_file)

    # Print quick summary
    print()
    locs = sorted(df['Lesson Location'].apply(normalize_location).unique())
    for loc in locs:
        n = len(df[df['Lesson Location'].apply(normalize_location) == loc])
        print(f'  Room {loc}: {n} lessons')


if __name__ == '__main__':
    main()

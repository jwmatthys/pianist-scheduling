"""
Accompanist Scheduler
=====================
Assigns piano accompanists to lessons based on availability, preference,
workload limits, block scheduling, required accompanist constraints,
and overlap handling.

Usage:
    python generate_pianist_schedule.py --workbook scheduling.xlsx

Workbook structure:
    - Sheet named "Lessons"        — the lesson schedule
    - Sheets prefixed "Pianist - " — one per accompanist (e.g. "Pianist - Michael")
    - Output written as new timestamped sheet: "Schedule 2026-02-20 14:35"

Lesson schedule columns:
    Lesson Teacher Name, Student Name, Lesson Day, Lesson Start Time,
    Lesson End Time, Lesson Location, Instrument, Required Accompanist (optional)

Availability values in each accompanist sheet:
    Available   - fully available (highest preference)
    Tentative   - available but lower preference
    Unavailable - cannot be assigned

Fit levels (highest to lowest):
    FULL     - all half-hour slots the lesson spans are "Available"
    PARTIAL  - all slots are "Available" or "Tentative" (≥1 Tentative)
    NEAR     - lesson start is within 15 min before an available/tentative boundary
    OVERLAP  - lesson overlaps ≤30 min with another lesson already assigned to this
               accompanist, and the accompanist covers the full combined window;
               only used when neither lesson has a required accompanist
    NONE     - no fit at all; best partial match assigned and flagged

Required accompanist:
    If "Required Accompanist" column is populated, only that accompanist may be
    assigned. If they are Unavailable for that time, the lesson is left unassigned
    and flagged. If they are a near/partial fit, they are assigned and flagged.

Workload:
    Each accompanist specifies a maximum weekly hours cap. Lessons are distributed
    as evenly as possible within caps. Over-cap assignments are flagged.

Block scheduling:
    Minimizes number of separate blocks and total gap time per accompanist per day.
"""

import argparse
from datetime import datetime, timedelta, time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS_ORDER              = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
NEAR_FIT_MARGIN_MINUTES = 15
OVERLAP_MAX_MINUTES     = 30

FIT_FULL    = 4
FIT_PARTIAL = 3
FIT_NEAR    = 2
FIT_OVERLAP = 1
FIT_NONE    = 0

FIT_LABELS = {FIT_FULL: "Full", FIT_PARTIAL: "Partial", FIT_NEAR: "Near",
              FIT_OVERLAP: "Overlap", FIT_NONE: "None"}
FIT_COLORS = {"Full": "C6EFCE", "Partial": "FFEB9C", "Near": "FFCC99",
              "Overlap": "DAE8FC", "None": "FFC7CE"}


# ── Time helpers ──────────────────────────────────────────────────────────────

def parse_time(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        h, rem = divmod(total, 3600)
        m = (rem // 60) % 60
        return time(h % 24, m)
    if isinstance(val, str):
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
            try:
                return datetime.strptime(val.strip(), fmt).time()
            except ValueError:
                pass
    raise ValueError(f"Cannot parse time: {val!r}")


def to_min(t):
    return t.hour * 60 + t.minute


def duration_hours(start, end):
    return (to_min(end) - to_min(start)) / 60.0


# ── Load accompanist sheet ────────────────────────────────────────────────────

LESSONS_SHEET   = "Lessons"
PIANIST_PREFIX  = "Pianist - "


def load_availability(df, sheet_name):
    """
    Parse an accompanist availability DataFrame (read with header=None).
    sheet_name is used as a fallback name if no name cell is found.
    Returns: name (str), max_hours (float|None), avail (dict)
        avail = {day: {slot_start_minutes: status}}
    """
    # Name — look for "Name:" label then take adjacent cell;
    # fall back to the part of the sheet name after "Pianist - "
    name = None
    for row_i in range(min(3, len(df))):
        row = df.iloc[row_i]
        for col in range(len(row)):
            val = row.iloc[col]
            if isinstance(val, str) and val.strip().lower().rstrip(":") == "name":
                for col2 in range(col + 1, len(row)):
                    v2 = row.iloc[col2]
                    if isinstance(v2, str) and v2.strip() \
                            and v2.strip().lower() not in ("your name here", "name"):
                        name = v2.strip()
                        break
                if name:
                    break
        if name:
            break
    if name is None:
        # Strip the "Pianist - " prefix from the sheet name
        name = sheet_name[len(PIANIST_PREFIX):].strip() if sheet_name.startswith(PIANIST_PREFIX) else sheet_name

    # Max hours
    max_hours = None
    for row_i in range(min(5, len(df))):
        row = df.iloc[row_i]
        for col in range(len(row)):
            val = row.iloc[col]
            if isinstance(val, str) and "hour" in val.lower():
                for col2 in range(col + 1, len(row)):
                    v2 = row.iloc[col2]
                    if pd.notna(v2):
                        try:
                            max_hours = float(v2)
                            break
                        except (ValueError, TypeError):
                            pass
                break
        if max_hours is not None:
            break

    # Day header row
    day_row_idx = None
    day_col_map = {}
    for i, row in df.iterrows():
        for j, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() in DAYS_ORDER:
                day_row_idx = i
                break
        if day_row_idx is not None:
            break
    if day_row_idx is None:
        raise ValueError(f"Could not find day header row in sheet '{sheet_name}'")
    for j, cell in enumerate(df.iloc[day_row_idx]):
        if isinstance(cell, str) and cell.strip() in DAYS_ORDER:
            day_col_map[j] = cell.strip()

    # Availability slots
    avail = {day: {} for day in day_col_map.values()}
    valid_statuses = {"Available", "Tentative", "Unavailable"}
    for i in range(day_row_idx + 1, len(df)):
        row = df.iloc[i]
        raw_time = row.iloc[0]
        if pd.isna(raw_time):
            continue
        try:
            t = parse_time(raw_time)
        except (ValueError, TypeError):
            continue
        slot_min = to_min(t)
        for j, day in day_col_map.items():
            status = row.iloc[j] if j < len(row) else None
            if isinstance(status, str) and status.strip() in valid_statuses:
                avail[day][slot_min] = status.strip()

    return name, max_hours, avail


# ── Fit scoring ───────────────────────────────────────────────────────────────

def get_fit(lesson_day, lesson_start, lesson_end, avail):
    """
    Returns (fit_score, has_tentative) based on the accompanist's availability.
    Does NOT consider overlap — that is handled separately in the assignment loop.
    """
    day_avail = avail.get(lesson_day, {})
    if not day_avail:
        return FIT_NONE, False

    start_min = to_min(lesson_start)
    end_min   = to_min(lesson_end)

    # Slots overlapping the lesson window (each slot covers 30 min)
    covered = [
        (sm, st) for sm, st in day_avail.items()
        if sm < end_min and sm + 30 > start_min
    ]

    if covered:
        statuses = [st for _, st in covered]
        if all(s == "Available" for s in statuses):
            return FIT_FULL, False
        if all(s in ("Available", "Tentative") for s in statuses):
            return FIT_PARTIAL, True

    # Near fit: lesson start within 15 min before an available/tentative boundary
    for slot_min, status in day_avail.items():
        if status in ("Available", "Tentative"):
            diff = slot_min - start_min
            if 0 <= diff <= NEAR_FIT_MARGIN_MINUTES:
                return FIT_NEAR, (status == "Tentative")

    return FIT_NONE, False


def get_fit_for_window(day, start_min, end_min, avail):
    """Check fit for an arbitrary minute window (used for overlap combined window check)."""
    day_avail = avail.get(day, {})
    covered = [
        st for sm, st in day_avail.items()
        if sm < end_min and sm + 30 > start_min
    ]
    if not covered:
        return FIT_NONE
    if all(s == "Available" for s in covered):
        return FIT_FULL
    if all(s in ("Available", "Tentative") for s in covered):
        return FIT_PARTIAL
    return FIT_NONE


def overlap_minutes(a_start, a_end, b_start, b_end):
    """Return the number of overlapping minutes between two windows."""
    return max(0, min(a_end, b_end) - max(a_start, b_start))


# ── Scatter penalty ───────────────────────────────────────────────────────────

def scatter_penalty(assignments):
    """
    assignments: list of (day, start_min, end_min)
    Returns blocks + total_gap_hours (equally weighted, lower = better).
    """
    if not assignments:
        return 0.0
    by_day = {}
    for day, s, e in assignments:
        by_day.setdefault(day, []).append((s, e))
    penalty = 0.0
    for slots in by_day.values():
        slots.sort()
        blocks, gap_min = 1, 0
        for i in range(1, len(slots)):
            gap = slots[i][0] - slots[i - 1][1]
            if gap > 0:
                blocks += 1
                gap_min += gap
        penalty += blocks + gap_min / 60.0
    return penalty


def resolve_required_name(raw, acc_names):
    """
    Match a raw 'Required Pianist' value against the list of loaded accompanist names.
    Returns the full matched name, or None if no match found.
    Matching strategy (in order):
      1. Exact match
      2. Case-insensitive exact match
      3. Any accompanist name that contains the raw value as a substring (case-insensitive)
      4. The raw value contains any accompanist name as a substring (case-insensitive)
    """
    if not raw:
        return None
    raw_l = raw.strip().lower()
    for name in acc_names:
        if name == raw.strip():
            return name
    for name in acc_names:
        if name.lower() == raw_l:
            return name
    for name in acc_names:
        if raw_l in name.lower():
            return name
    for name in acc_names:
        if name.lower() in raw_l:
            return name
    return None


# ── Core assignment ───────────────────────────────────────────────────────────

def assign_lessons(lessons_df, accompanists):
    """
    Main assignment loop.
    accompanists: list of (name, max_hours, avail_dict)

    Assignment priority per lesson:
      1. Required accompanist (hard constraint)
      2. Best standard fit (Full > Partial > Near)
      3. Overlap fit (only when neither lesson has a required accompanist)
      4. Conflict (no fit — best partial match)

    Within each fit tier: prefer not-over-cap > non-tentative > lower workload > less scatter.

    Returns: (list of result dicts, dict of hours_per_accompanist)
    """
    day_order = {d: i for i, d in enumerate(DAYS_ORDER)}
    df = lessons_df.copy()

    # Filter to rows where Need Pianist is checked (True / 1 / "TRUE")
    need_col = next((c for c in df.columns if c.strip().lower() == "need pianist"), None)
    if need_col:
        df = df[df[need_col].apply(lambda v: str(v).strip().upper() in ("1", "1.0", "TRUE", "YES"))]
        df = df.reset_index(drop=True)

    # Drop rows with no valid day or time (empty trailing rows)
    df = df[df["Lesson Day"].notna() & df["Lesson Start Time"].notna()].reset_index(drop=True)

    df["_day_order"] = df["Lesson Day"].map(lambda d: day_order.get(d, 99))
    df["_start_min"] = df["Lesson Start Time"].map(lambda t: to_min(parse_time(t)))
    df["_end_min"]   = df["Lesson End Time"].map(lambda t: to_min(parse_time(t)))
    df = df.sort_values(["_day_order", "_start_min"]).reset_index(drop=True)

    # Find required pianist column (accepts "Required Pianist" or "Required Accompanist")
    req_col = next((c for c in df.columns
                    if "required" in c.lower() and ("pianist" in c.lower() or "accompanist" in c.lower())), None)

    acc_names         = [name for name, _, _ in accompanists]
    acc_map           = {name: avail for name, _, avail in accompanists}
    current_assigns   = {name: [] for name, _, _ in accompanists}
    current_hours     = {name: 0.0  for name, _, _ in accompanists}
    max_hours_map     = {name: mh   for name, mh, _ in accompanists}
    results           = []

    for idx, lesson in df.iterrows():
        day       = lesson["Lesson Day"]
        start     = parse_time(lesson["Lesson Start Time"])
        end       = parse_time(lesson["Lesson End Time"])
        dur       = duration_hours(start, end)
        s_min     = to_min(start)
        e_min     = to_min(end)

        # Resolve required pianist via partial matching
        required_raw  = None
        required      = None   # fully resolved name
        if req_col and pd.notna(lesson.get(req_col, None)):
            required_raw = str(lesson[req_col]).strip() or None
        if required_raw:
            required = resolve_required_name(required_raw, acc_names)

        flags         = []
        assigned_name = None
        fit_score     = FIT_NONE
        has_tentative = False
        overlap_pair  = None

        # ── CASE 1: Required accompanist ─────────────────────────────────────
        if required_raw:
            if required is None:
                flags.append(f"⚠ REQUIRED PIANIST '{required_raw}' NOT FOUND — lesson unassigned")
            else:
                fit_score, has_tentative = get_fit(day, start, end, acc_map[required])
                if fit_score == FIT_NONE:
                    flags.append(f"⚠ REQUIRED PIANIST '{required}' is UNAVAILABLE — lesson unassigned")
                    assigned_name = None
                else:
                    assigned_name = required
                    if fit_score == FIT_NEAR:
                        flags.append(f"⚠ NEAR FIT: '{required}' available within 15 min of lesson start")
                    elif fit_score == FIT_PARTIAL:
                        flags.append(f"ℹ PARTIAL FIT: '{required}' has Tentative availability for this slot")
                    if has_tentative:
                        flags.append("ℹ TENTATIVE availability")
                    mh = max_hours_map.get(required)
                    if mh and current_hours[required] + dur > mh:
                        flags.append(f"⚠ OVER CAP: Exceeds {mh}h weekly limit")
                    # Note if we matched on a partial name
                    if required_raw.strip().lower() != required.lower():
                        flags.append(f"ℹ Required pianist '{required_raw}' matched to '{required}'")

        # ── CASE 2: Normal assignment ─────────────────────────────────────────
        else:
            candidates = []
            for name, _, avail in accompanists:
                fit, tentative = get_fit(day, start, end, avail)
                hours_after   = current_hours[name] + dur
                mh            = max_hours_map.get(name)
                over_cap      = bool(mh and hours_after > mh)
                workload      = (hours_after / mh) if mh else hours_after
                new_assigns   = current_assigns[name] + [(day, s_min, e_min)]
                scat          = scatter_penalty(new_assigns)
                candidates.append({
                    "name": name, "fit": fit, "tentative": tentative,
                    "over_cap": over_cap, "workload": workload, "scatter": scat,
                    "hours_after": hours_after, "max_hours": mh,
                })

            # Filter to best standard fit tier (Full/Partial/Near)
            best_standard_fit = max(c["fit"] for c in candidates)

            if best_standard_fit >= FIT_NEAR:
                # Use standard fit candidates
                pool = [c for c in candidates if c["fit"] == best_standard_fit]
                pool.sort(key=lambda c: (c["over_cap"], c["tentative"], c["workload"], c["scatter"]))
                chosen       = pool[0]
                assigned_name = chosen["name"]
                fit_score     = chosen["fit"]
                has_tentative = chosen["tentative"]

            else:
                # ── Try overlap fit ───────────────────────────────────────────
                # Find already-assigned lessons on the same day that overlap ≤30 min
                # with this lesson, where the paired lesson had NO required accompanist.
                overlap_candidate = None
                for prev in results:
                    if prev["Lesson Day"] != day:
                        continue
                    if prev.get("_required"):      # skip if that lesson had a required acc
                        continue
                    prev_s = to_min(parse_time(prev["Lesson Start Time"]))
                    prev_e = to_min(parse_time(prev["Lesson End Time"]))
                    ov = overlap_minutes(s_min, e_min, prev_s, prev_e)
                    if 0 < ov <= OVERLAP_MAX_MINUTES:
                        # Check if the paired accompanist covers the full combined window
                        paired_name  = prev["Assigned Accompanist"]
                        combined_s   = min(s_min, prev_s)
                        combined_e   = max(e_min, prev_e)
                        combined_fit = get_fit_for_window(
                            day, combined_s, combined_e, acc_map[paired_name])
                        if combined_fit >= FIT_PARTIAL:
                            mh        = max_hours_map.get(paired_name)
                            over_cap  = bool(mh and current_hours[paired_name] + dur > mh)
                            workload  = ((current_hours[paired_name] + dur) / mh) if mh else (current_hours[paired_name] + dur)
                            new_a     = current_assigns[paired_name] + [(day, s_min, e_min)]
                            scat      = scatter_penalty(new_a)
                            overlap_candidate = {
                                "name": paired_name, "fit": FIT_OVERLAP,
                                "tentative": False, "over_cap": over_cap,
                                "workload": workload, "scatter": scat,
                                "paired_with": prev,
                            }
                            break   # take the first valid overlap partner found

                if overlap_candidate:
                    assigned_name = overlap_candidate["name"]
                    fit_score     = FIT_OVERLAP
                    has_tentative = False
                    overlap_pair  = overlap_candidate["paired_with"]
                    prev_student  = overlap_pair["Student"]
                    flags.append(f"ℹ OVERLAP FIT: shares accompanist with {prev_student} (lessons overlap ≤30 min)")
                    if overlap_candidate["over_cap"]:
                        mh = max_hours_map.get(assigned_name)
                        flags.append(f"⚠ OVER CAP: Exceeds {mh}h weekly limit")
                    # Also annotate the earlier lesson
                    overlap_pair["Notes"] += f" | ℹ OVERLAP FIT: paired with {lesson.get('Student Name', '')}"
                    overlap_pair["Fit Quality"] = FIT_LABELS[FIT_OVERLAP]

                else:
                    # ── Conflict: assign best partial match ───────────────────
                    candidates.sort(key=lambda c: (c["over_cap"], c["workload"], c["scatter"]))
                    chosen        = candidates[0]
                    assigned_name = chosen["name"]
                    fit_score     = FIT_NONE
                    flags.append("⚠ CONFLICT: No available accompanist — best match assigned")
                    if chosen["over_cap"]:
                        flags.append(f"⚠ OVER CAP: Exceeds {chosen['max_hours']}h weekly limit")

            # Standard-fit flags
            if fit_score == FIT_NEAR:
                flags.append("⚠ NEAR FIT: Lesson start within 15 min of available window")
            if has_tentative and fit_score not in (FIT_NONE, FIT_OVERLAP):
                flags.append("ℹ TENTATIVE availability")
            if fit_score >= FIT_NEAR and assigned_name:
                chosen_c = next((c for c in candidates if c["name"] == assigned_name), None)
                if chosen_c and chosen_c["over_cap"]:
                    flags.append(f"⚠ OVER CAP: Exceeds {chosen_c['max_hours']}h weekly limit")

        # ── Record result ─────────────────────────────────────────────────────
        if assigned_name:
            current_assigns[assigned_name].append((day, s_min, e_min))
            current_hours[assigned_name] += dur

        results.append({
            "Lesson Day":           day,
            "Lesson Start Time":    lesson["Lesson Start Time"],
            "Lesson End Time":      lesson["Lesson End Time"],
            "Teacher":              lesson.get("Lesson Teacher Name", ""),
            "Student":              lesson.get("Student Name", ""),
            "Instrument":           lesson.get("Instrument", ""),
            "Location":             lesson.get("Lesson Location", ""),
            "Required Pianist":     required_raw or "",
            "Assigned Accompanist": assigned_name or "UNASSIGNED",
            "Hours":                round(dur, 2),
            "Fit Quality":          FIT_LABELS[fit_score],
            "Notes":                " | ".join(flags),
            "_required":            bool(required_raw),
        })

    return results, current_hours


# ── Excel output ──────────────────────────────────────────────────────────────

def write_output(results, hours_summary, accompanists_meta, workbook_path):
    from datetime import datetime as dt
    from pathlib import Path

    timestamp        = dt.now().strftime("%Y-%m-%d %H:%M")
    file_timestamp   = dt.now().strftime("%Y-%m-%d_%H-%M")
    assignments_name = "Schedule - Assignments"
    weekly_name      = "Schedule - Weekly"
    summary_name     = "Schedule - Summary"

    # Build output filename: original_stem + timestamp + .xlsx
    p           = Path(workbook_path)
    output_path = p.parent / f"{p.stem}_{file_timestamp}.xlsx"

    # Create a fresh workbook for output only
    import shutil
    wb = Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    thin = Side(style="thin", color="CCCCCC")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, value, bg="1F4E79"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return c

    def dat(ws, row, col, value, bg=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="center")
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        return c

    def add_timestamp_row(ws, ncols):
        """Write a 'Generated: <timestamp>' banner across the top of a sheet."""
        c = ws.cell(row=1, column=1, value=f"Generated: {timestamp}")
        c.font      = Font(italic=True, color="595959", name="Arial", size=9)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.row_dimensions[1].height = 16
        return 2   # return the next available row

    max_hours_map = {name: mh for name, mh, _ in accompanists_meta}

    # ── Sheet 1: Assignments ──────────────────────────────────────────────────
    ws = wb.create_sheet(assignments_name)

    col_headers = ["Day", "Start", "End", "Teacher", "Student", "Instrument",
                   "Location", "Required", "Accompanist", "Hours", "Fit", "Notes"]
    col_keys    = ["Lesson Day", "Lesson Start Time", "Lesson End Time", "Teacher",
                   "Student", "Instrument", "Location", "Required Pianist",
                   "Assigned Accompanist", "Hours", "Fit Quality", "Notes"]

    data_start = add_timestamp_row(ws, len(col_headers))

    for col, h in enumerate(col_headers, 1):
        hdr(ws, data_start, col, h)

    for row_idx, r in enumerate(results, data_start + 1):
        for col_idx, key in enumerate(col_keys, 1):
            val = r[key]
            if hasattr(val, "hour"):
                val = val.strftime("%H:%M")
            dat(ws, row_idx, col_idx, str(val) if val else "")

        fit_val = r["Fit Quality"]
        ws.cell(row=row_idx, column=11).fill = PatternFill(
            "solid", start_color=FIT_COLORS.get(fit_val, "FFFFFF"))
        ws.cell(row=row_idx, column=11).alignment = Alignment(horizontal="center")

        notes = r["Notes"]
        if "UNASSIGNED" in r["Assigned Accompanist"] or "CONFLICT" in notes or "OVER CAP" in notes:
            row_bg = "FCE4D6"
        elif "NEAR FIT" in notes or "TENTATIVE" in notes or "OVERLAP" in notes:
            row_bg = "FFF2CC"
        else:
            row_bg = None
        if row_bg:
            for ci in range(1, 11):
                ws.cell(row=row_idx, column=ci).fill = PatternFill("solid", start_color=row_bg)

    for i, w in enumerate([11, 7, 7, 14, 14, 12, 10, 14, 20, 6, 8, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[data_start].height = 30
    ws.freeze_panes = f"A{data_start + 1}"

    # ── Sheet 2: Weekly View ──────────────────────────────────────────────────
    ws2 = wb.create_sheet(weekly_name)
    ws2.sheet_view.showGridLines = False

    acc_names = sorted(set(r["Assigned Accompanist"] for r in results
                           if r["Assigned Accompanist"] != "UNASSIGNED"))
    grid_data = {name: {day: [] for day in DAYS_ORDER} for name in acc_names}
    for r in results:
        if r["Assigned Accompanist"] != "UNASSIGNED":
            grid_data[r["Assigned Accompanist"]][r["Lesson Day"]].append(r)

    days_used = sorted(
        set(r["Lesson Day"] for r in results),
        key=lambda d: DAYS_ORDER.index(d) if d in DAYS_ORDER else 99
    )
    for col in range(1, len(days_used) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 24

    current_row = add_timestamp_row(ws2, max(len(days_used), 1))
    for name in acc_names:
        cap_str = f" / {max_hours_map[name]}h cap" if max_hours_map.get(name) else ""
        c = ws2.cell(row=current_row, column=1,
                     value=f"{name}  —  {round(hours_summary.get(name, 0), 1)}h assigned{cap_str}")
        c.font = Font(bold=True, name="Arial", size=11, color="1F4E79")
        current_row += 1

        for col, day in enumerate(days_used, 1):
            hdr(ws2, current_row, col, day, bg="2E75B6")
        current_row += 1

        day_lessons = {
            day: sorted(grid_data[name][day],
                        key=lambda r: to_min(parse_time(r["Lesson Start Time"])))
            for day in days_used
        }
        max_slots = max((len(v) for v in day_lessons.values()), default=0)

        for slot in range(max_slots):
            ws2.row_dimensions[current_row].height = 50
            for col, day in enumerate(days_used, 1):
                day_list = day_lessons[day]
                if slot < len(day_list):
                    r  = day_list[slot]
                    s  = parse_time(r["Lesson Start Time"]).strftime("%H:%M")
                    e  = parse_time(r["Lesson End Time"]).strftime("%H:%M")
                    req_str = f" [REQ]" if r["Required Pianist"] else ""
                    text = f"{s}–{e}{req_str}\n{r['Teacher']} / {r['Student']}"
                    if r["Notes"]:
                        text += f"\n{r['Notes'][:45]}"
                    c = ws2.cell(row=current_row, column=col, value=text)
                    c.font      = Font(name="Arial", size=9)
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                    c.border    = bord
                    c.fill      = PatternFill("solid", start_color=FIT_COLORS.get(r["Fit Quality"], "FFFFFF"))
                else:
                    ws2.cell(row=current_row, column=col, value="").border = bord
            current_row += 1

        current_row += 2

    # Unassigned lessons section
    unassigned = [r for r in results if r["Assigned Accompanist"] == "UNASSIGNED"]
    if unassigned:
        c = ws2.cell(row=current_row, column=1, value="⚠ UNASSIGNED LESSONS")
        c.font = Font(bold=True, name="Arial", size=11, color="FF0000")
        current_row += 1
        for r in unassigned:
            s = parse_time(r["Lesson Start Time"]).strftime("%H:%M")
            e = parse_time(r["Lesson End Time"]).strftime("%H:%M")
            c = ws2.cell(row=current_row, column=1,
                         value=f"{r['Lesson Day']}  {s}–{e}  {r['Teacher']} / {r['Student']}  — {r['Notes']}")
            c.font = Font(name="Arial", size=10, color="CC0000")
            current_row += 1

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet(summary_name)
    sum_headers = ["Accompanist", "Max Hours", "Assigned Hours", "% of Cap",
                   "Lessons", "Full", "Partial", "Near", "Overlap", "Conflicts",
                   "Over Cap", "Unassigned (required)"]
    s3_data_start = add_timestamp_row(ws3, len(sum_headers))
    for col, h in enumerate(sum_headers, 1):
        hdr(ws3, s3_data_start, col, h)

    all_names = sorted(set(r["Assigned Accompanist"] for r in results))
    for row_i, name in enumerate(all_names, s3_data_start + 1):
        person   = [r for r in results if r["Assigned Accompanist"] == name]
        assigned = round(hours_summary.get(name, 0), 2)
        max_h    = max_hours_map.get(name)
        pct      = f"{round(assigned / max_h * 100, 1)}%" if max_h else "N/A"

        vals = [
            name, max_h or "None", assigned, pct,
            len(person),
            sum(1 for r in person if r["Fit Quality"] == "Full"),
            sum(1 for r in person if r["Fit Quality"] == "Partial"),
            sum(1 for r in person if r["Fit Quality"] == "Near"),
            sum(1 for r in person if r["Fit Quality"] == "Overlap"),
            sum(1 for r in person if r["Fit Quality"] == "None"),
            sum(1 for r in person if "OVER CAP" in r["Notes"]),
            sum(1 for r in results if r["Assigned Accompanist"] == "UNASSIGNED"
                and resolve_required_name(r.get("Required Pianist", ""), [name]) is not None),
        ]
        for col, val in enumerate(vals, 1):
            c = dat(ws3, row_i, col, val)
            if col == 4 and max_h and assigned > max_h:
                c.fill = PatternFill("solid", start_color="FFC7CE")
        ws3.row_dimensions[row_i].height = 18

    for col in range(1, len(sum_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 15
    ws3.column_dimensions["A"].width = 22
    ws3.row_dimensions[s3_data_start].height = 30
    ws3.freeze_panes = f"A{s3_data_start + 1}"

    wb.save(output_path)
    print(f"\nOutput written to '{output_path}'")
    print(f"  Sheets added: '{assignments_name}', '{weekly_name}', '{summary_name}'")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Assign accompanists to lessons.")
    parser.add_argument("--workbook", required=True,
                        help="Excel workbook containing Lessons sheet and Pianist - * sheets")
    args = parser.parse_args()

    print(f"Loading workbook: {args.workbook}")
    xl = pd.ExcelFile(args.workbook)

    # Load lessons sheet
    if LESSONS_SHEET not in xl.sheet_names:
        raise ValueError(f"No sheet named '{LESSONS_SHEET}' found in workbook. "
                         f"Available sheets: {xl.sheet_names}")
    lessons_df = xl.parse(LESSONS_SHEET)
    print(f"  '{LESSONS_SHEET}': {len(lessons_df)} lessons found")

    # Load all accompanist sheets
    pianist_sheets = [s for s in xl.sheet_names if s.startswith(PIANIST_PREFIX)]
    if not pianist_sheets:
        raise ValueError(f"No sheets starting with '{PIANIST_PREFIX}' found in workbook.")

    accompanists = []
    for sheet_name in pianist_sheets:
        df = xl.parse(sheet_name, header=None)
        name, max_hours, avail = load_availability(df, sheet_name)
        accompanists.append((name, max_hours, avail))
        cap_str = f"{max_hours}h cap" if max_hours else "no cap"
        print(f"  '{sheet_name}': loaded as '{name}' ({cap_str})")

    print("\nRunning assignment algorithm...")
    results, hours_summary = assign_lessons(lessons_df, accompanists)

    # Strip internal keys before reporting
    for r in results:
        r.pop("_required", None)

    total      = len(results)
    unassigned = [r for r in results if r["Assigned Accompanist"] == "UNASSIGNED"]
    print(f"\nResults ({total} lessons):")
    print(f"  Full fits:    {sum(1 for r in results if r['Fit Quality'] == 'Full')}")
    print(f"  Partial fits: {sum(1 for r in results if r['Fit Quality'] == 'Partial')}")
    print(f"  Near fits:    {sum(1 for r in results if r['Fit Quality'] == 'Near')}")
    print(f"  Overlap fits: {sum(1 for r in results if r['Fit Quality'] == 'Overlap')}")
    print(f"  Conflicts:    {sum(1 for r in results if r['Fit Quality'] == 'None')}")
    print(f"  Unassigned:   {len(unassigned)}")
    print(f"  Over cap:     {sum(1 for r in results if 'OVER CAP' in r['Notes'])}")

    print("\nWorkload summary:")
    for name, max_h, _ in accompanists:
        h    = round(hours_summary.get(name, 0), 2)
        cap  = f"/ {max_h}h cap" if max_h else "(no cap)"
        flag = " ⚠ OVER CAP" if max_h and h > max_h else ""
        print(f"  {name}: {h}h {cap}{flag}")

    if unassigned:
        print("\n⚠ UNASSIGNED LESSONS (required accompanist unavailable):")
        for r in unassigned:
            s = parse_time(r["Lesson Start Time"]).strftime("%H:%M")
            print(f"  {r['Lesson Day']} {s} — {r['Teacher']} / {r['Student']}"
                  f"  (required: {r['Required Pianist']})")

    write_output(results, hours_summary, accompanists, args.workbook)


if __name__ == "__main__":
    main()
